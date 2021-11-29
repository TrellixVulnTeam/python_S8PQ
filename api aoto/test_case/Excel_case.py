import xlrd
import openpyxl
# #获取路径
# xlPath = 'C:\\Users\\张铁瀛\\PycharmProjects\\api aoto\\test_data\\case.xlsx'
# #读取Excel
# xlBook = xlrd.open_workbook(xlPath)
# #读取工作簿数量
# count = len(xlBook.sheets())
# print("工作簿数量为：", count)
# table = xlBook.sheets()[0]
# nrows = table.nrows
# ncols = table.ncols
# print("表数据行列为(%d, %d)" % (nrows, ncols))
# for i in range(0, nrows):
#     rowValues = table.row_values(i)
#     for data in rowValues:
#         print(data, " ")
#     print("")

workbook = openpyxl.load_workbook('C:\\Users\\张铁瀛\\PycharmProjects\\api aoto\\test_data\\case.xlsx')
sheet = workbook['Sheet1']
cell = sheet.cell(row=1, column=1).value
print(cell)
#获取最大行数
print(sheet.max_row)
print(sheet.max_column)
rows = sheet.rows
for i in list(rows):
    case = []
    for c in i:
        case.append(c.value)
    print(case)